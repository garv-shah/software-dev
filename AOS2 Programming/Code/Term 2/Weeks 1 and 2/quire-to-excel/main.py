import csv
from datetime import datetime
from openpyxl import load_workbook
from copy import copy
import urllib.request


def insert_row(row_num, copy_from):
    gantt.insert_rows(row_num)
    for row in gantt.iter_rows(min_row=copy_from, max_col=128, max_row=copy_from):
        for cell in row:
            new_cell = gantt.cell(row=row_num, column=cell.column, value=cell.value)
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)


wb = load_workbook('gantt-template.xlsx')
gantt = wb['ProjectSchedule']

# load csv
quire = []
with open("input.csv", 'r') as csv_file:
    # Creates a dictionary from the csv file
    reader = csv.DictReader(csv_file)
    quire = list(reader)
    # Sets the values of the layout fields to the values in the csv file
    csv_file.close()

milestones = []
tasks = {}
current_milestone = 'nil'

for task in quire:
    if task['\ufeff"ID"'] == '':
        tasks[current_milestone].append(task)
    else:
        milestones.append(task)
        current_milestone = task['Name']
        tasks[current_milestone] = [task]

for i in range(7):
    gantt.cell(row=8 + 2 * i, column=2).value = milestones[i]['Name']

row_num = 10
for i in range(len(tasks.keys()) - 1):
    subtasks = tasks[milestones[i]['Name']]
    print(milestones[i]['Name'])
    print(len(subtasks))
    print(9 + 2 * i)
    task_num = 0

    for x in range(len(subtasks)):
        num = row_num + 1 * i
        if x + 1 != len(subtasks):
            insert_row(num, num - 1)
        gantt.cell(column=2, row=num - 1).value = subtasks[task_num]['Name']
        gantt.cell(column=3, row=num - 1).value = subtasks[task_num]['Assignee']

        if subtasks[task_num]['Status'] == 'In Progress':
            gantt.cell(column=4, row=num - 1).value = 0.5
        elif subtasks[task_num]['Status'] == 'Completed':
            gantt.cell(column=4, row=num - 1).value = 1

        try:
            gantt.cell(column=5, row=num - 1).value = datetime.strptime(subtasks[task_num]['Start'], '%d %b %Y').date()
        except ValueError:
            print('no starting date')

        try:
            gantt.cell(column=6, row=num - 1).value = datetime.strptime(subtasks[task_num]['Due'], '%d %b %Y').date()
        except ValueError:
            print('no ending date')

        row_num += 1
        task_num += 1
    # row_num += 2

print(wb['ProjectSchedule']['B1'].value)
wb.save('SAT Gantt Chart.xlsx')
